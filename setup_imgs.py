import os
import time
import sqlite3
import openpyxl
import requests
from bs4 import BeautifulSoup
from settings import *


class Infomation(BeautifulSoup):
    def __init__(self, html_context: str, title: str):
        super().__init__(html_context, "html5lib")
        self.title = title  # 信息标题
        if not os.path.exists(file := "D:/medical_images/ct_infomations/{}".format(title)):
            os.mkdir(file)
        self.filepath = file + "/"  # 本地真实路径

    def completeness_check(self):
        """
        数据完整性检测
        """
        result = {"Legend which related existed": True}
        # 检测是否能关联到legend
        with sqlite3.connect(DB_PATH) as db:
            cur = db.cursor()
            cur.execute(
                "SELECT COUNT(*) FROM mi_legends WHERE en_name LIKE ? OR infomation LIKE ?",
                (f"%{self.title}%", f"%{self.title}%")
            )
            if cur.fetchone()[0] == 0:
                result["Legend which related existed"] = False
        return result

    def save_images(self):
        """
        
        """
        for i, img in enumerate(self.find_all("img")):
            if not self.image_right_size(img.attrs):
                continue
            url = str(img.attrs["src"])
            try:
                extend = url[url.rindex("."):]
                assert len(extend) < 5
            except (ValueError, AssertionError):
                extend = ".jpg"
            filename = "{}_{}{}".format(self.title, i, extend)

            if not os.path.exists(self.filepath + filename):
                if not url.startswith("http"):
                    url = "https:{}".format(url)
                try:
                    if 'wikipedia' in url:
                        data = requests.get(url, headers=WIKI_HEADERS)
                    else:
                        data = requests.get(url)
                    print("{} downloaded".format(filename), data.status_code)
                    if data.status_code < 400:
                        with open(self.filepath + filename, 'wb') as f:
                            for chunk in data.iter_content():
                                f.write(chunk)
                except requests.RequestException as e:
                    print("Image download failed:", e)
                finally:
                    time.sleep(1)

            img.attrs["src"] = "{}{}/{}".format(self.absolute_path, self.title, filename)

    def save_html(self):
        """
        保存修改当富文本
        """
        if not os.path.exists(newpath := self.filepath + self.title + ".html"):
            with open(newpath, 'w', encoding='UTF-8') as f:
                f.write(self.prettify(formatter='html'))

    def change_legends(self):
        """
        改变legends数据infomaton地址
        """
        with sqlite3.connect(DB_PATH) as db:
            cur = db.cursor()
            cur.execute(
                "SELECT value, infomation FROM mi_legends WHERE infomation LIKE ? AND infomation NOT LIKE ?",
                (f"%{self.title}%", f"{self.absolute_path}%")
            )
            for value, origin_link in cur.fetchall():
                cur.execute(
                    "INSERT INTO mi_legend_infomation_history (legend_value, origin_link) VALUES (?,?)",
                    (value, origin_link)
                )
                cur.execute(
                    "UPDATE mi_legends SET infomation=? WHERE value=?",
                    ("{}{}/{}.html".format(self.absolute_path, self.title, self.title), value)
                )
                db.commit()

    @classmethod
    def read_local(cls, path: str):
        """
        构造类方法
        """
        with open(path, 'r', encoding='utf-8') as f:
            context = f.read()
            s = path.rfind("/")
            e = path.rfind(".")
            name = path[s + 1: e]
            return cls(context, name)

    @staticmethod
    def image_right_size(attrs: dict):
        """
        检测img标签合法性
        """
        if "width" in attrs and "height" in attrs:
            return int(attrs["width"]) > 100 or int(attrs["height"]) > 100
        elif "style" in attrs:
            w = h = 0
            for css in attrs['style'].split(';'):
                css = css.strip()
                try:
                    if css.startswith("height"):
                        w = int(css[css.index(":") + 1:css.index("px")])
                    elif css.startswith("width"):
                        h = int(css[css.index(":") + 1:css.index("px")])
                except ValueError:
                    w = h = 101
            return w > 100 or h > 100


def reduction():
    with sqlite3.connect(DB_PATH) as db:
        cur = db.cursor()
        cur.execute("SELECT value FROM mi_legends WHERE infomation IS NULL OR infomation=''")
        for value, in cur.fetchall():
            cur.execute("SELECT origin_link FROM mi_legend_infomation_history WHERE legend_value=?", (value,))
            _res = cur.fetchone()
            if _res is not None:
                cur.execute("UPDATE mi_legends SET infomation=? WHERE value=?", (_res[0], value))
                db.commit()


def test():
    db = sqlite3.connect(DB_PATH)
    cur = db.cursor()
    wb = openpyxl.load_workbook("project/lost.xlsx")
    ws = wb.active
    # for row in ws.iter_rows(values_only=True):
    #     cur.execute("SELECT COUNT(*) FROM mi_legends WHERE infomation LIKE ?", (f"%{row[1]}%",))
    #     if cur.fetchone()[0] == 0:
    #         cur.execute(
    #         "SELECT COUNT(*) FROM mi_legend_infomation_history WHERE origin_link LIKE ?", (f"%{row[1]}%",)
    #         )
    #         if cur.fetchone()[0] == 0:
    #             print(row)
    cur.execute("SELECT infomation FROM mi_legends WHERE infomation NOT LIKE ?", ("%xlhcq%",))
    for info, in cur.fetchall():
        for row in ws.iter_rows(values_only=True):
            if info in row[1] or row[1] in info:
                break
        else:
            print(info)


if __name__ == '__main__':
    pass
